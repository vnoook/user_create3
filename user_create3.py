# Скрипт создания файла новых пользователей
# для его работы нужно
# 1) чтобы инфа начиналась с 15 строки
# 2) не было невидимых строк и колонок
# 3) не было объединённых ячеек в списке фамилий
# 4) список подразделений желательно заменить на коды, либо это сделать уже в новом файле
# 5) в экселе надо выбирать "сохр как" и "CSV (разделите - запятые)"
# 6) если в папке скрипта будет файл xl_f_exists,
#    то существующие отпечатки и фио будут сравниваться с отпечатками и фио в заявке
# ...
# INSTALL
# pip install openpyxl
# ...

import os
import time
import openpyxl

# считаю время скрипта
print('начинается.......')
time_start = time.monotonic()

# словарь для соответствия русских символов и английских
alfa_dic = {'а': 'A', 'б': 'B', 'в': 'V', 'г': 'G', 'д': 'D', 'е': 'E', 'ё': 'E', 'ж': 'ZH', 'з': 'Z', 'и': 'I',
            'й': 'I', 'к': 'K', 'л': 'L', 'м': 'M', 'н': 'N', 'о': 'O', 'п': 'P', 'р': 'R', 'с': 'S', 'т': 'T',
            'у': 'U', 'ф': 'F', 'х': 'H', 'ц': 'TS', 'ч': 'CH', 'ш': 'SH', 'щ': 'SHCH', 'ъ': '', 'ы': 'Y', 'ь': '',
            'э': 'E', 'ю': 'IU', 'я': 'IA', 'А': 'A', 'Б': 'B', 'В': 'V', 'Г': 'G', 'Д': 'D', 'Е': 'E',
            'Ё': 'E', 'Ж': 'ZH', 'З': 'Z', 'И': 'I', 'Й': 'I', 'К': 'K', 'Л': 'L', 'М': 'M', 'Н': 'N', 'О': 'O',
            'П': 'P', 'Р': 'R', 'С': 'S', 'Т': 'T', 'У': 'U', 'Ф': 'F', 'Х': 'H', 'Ц': 'TS', 'Ч': 'CH', 'Ш': 'SH',
            'Щ': 'SHCH', 'Ъ': '', 'Ы': 'Y', 'Ь': '', 'Э': 'E', 'Ю': 'IU', 'Я': 'IA'}


# функция транслитерация
def latinizator(stroka, dic):
    for key, value in dic.items():
        stroka = stroka.replace(key, value)
    return stroka


# функция обработки отпечатка сертификата, убирает всё лишнее оставляя только цифры и буквы
def stripper(string_val):
    if string_val:
        strip_val = ''.join(char for char in string_val if char.isalnum())
    else:
        strip_val = ''
    return strip_val


# стартовая строка на которой начинается поиск на листе
string_first = 15
# конечная строка на листе
string_last = None
# длина отпечатка сертификата
otp_len = 40

# файл из которого беру заявку(и)
xl_f_from = 'users_info.xlsx'
# файл для новых юзеров
xl_f_to = 'users_new.xlsx'
# файл с существующими пользователями для проверки отпечатка и фио
xl_f_exists = 'users_exists.xlsx'
xl_f_exists_s = 'Пользователи'

# открываю книгу с заявками на пользователей
wb_from = openpyxl.load_workbook(xl_f_from)

# создаю итоговую книгу для экспорта
wb_new = openpyxl.Workbook()
wb_new_s = wb_new.active

# формирую первую строку в новом файле
wb_new_s.append(['Пользователь', 'Фамилия', 'Имя', 'Отчество', 'Должность', 'Пароль',
                 'Отпечаток сертификата', 'Подразделение', 'Роль'])

# Проверка на существование файла с отпечатками сертификатов и фио
# если файла нет, то проверка не производится
otp_set_exists = set()
fio_set_exists = set()
if os.path.exists(xl_f_exists):
    # открываю книгу с уже существующими пользователями
    wb_exists = openpyxl.load_workbook(xl_f_exists)
    wb_exists_s = wb_exists[xl_f_exists_s]

    # беру все отпечатки и фио создаю из них множества
    for xl_row in range(2, wb_exists_s.max_row + 1):
        fio_set_exists.add(''.join((wb_exists_s.cell(xl_row, 2).value,
                                    wb_exists_s.cell(xl_row, 3).value,
                                    wb_exists_s.cell(xl_row, 4).value)).lower())
        if wb_exists_s.cell(xl_row, 7).value:
            otp_set_exists.add(wb_exists_s.cell(xl_row, 7).value)

    wb_exists.close()
else:
    print()
    print('*'*50)
    print(f'Отсутствует файл {xl_f_exists} в текущей папке.')
    print(f'Он требуется для сверки отпечатков сертификатов и фио.')
    print('*'*50)
    print()

# список отпечатков в файле заявки, если их больше двух, то не добавлять повторные
otp_list = []

for xl_sheet in wb_from.sheetnames:
    # назначение активного листа
    wb_from_s = wb_from[xl_sheet]
    # вычисление последней строки
    string_last = wb_from_s.max_row

    # собираю информацию для заполнения конечной строки формата
    # Пользователь;Фамилия;Имя;Отчество;Должность;Пароль;Отпечаток сертификата;Подразделение;Роль
    # ---
    # прохожусь по строкам в активном листе
    for xl_row in range(string_first, string_last+1):
        # флаг добавления строки в файл экспорта
        flag_add_in_new_file = True

        # прохожусь по колонкам активного листа
        for xl_col in range(2, wb_from_s.max_column+1):
            # беру значение ячейки обрезая пробелы по бокам
            cell_value = str(wb_from_s.cell(xl_row, xl_col).value)
            # print(wb_from_s.cell(xl_row, xl_col).coordinate, xl_row, xl_col, type(cell_value), cell_value)
            if cell_value is not None:
                cell_value = cell_value.strip()

            # НОВЫЙ ПОЛЬЗОВАТЕЛЬ, ФОРМИРУЮ СТРОКУ
            # Пользователь Фамилия Имя Отчество
            if xl_col == 2:
                fio = cell_value.split()
                fam_cell = fio[0].strip()
                imya_cell = fio[1].strip()
                otch_cell = fio[2].replace(',', '').strip()

                # проверка на существование такого фио (пользователя)
                # если такого фио нет, то добавить в строку для выгрузки в файл
                if (''.join((fam_cell, imya_cell, otch_cell))).lower() in fio_set_exists:
                    flag_add_in_new_file = False
                else:
                    # тут делается пользователь английскими буквами
                    user_name = latinizator(fam_cell + '.' + imya_cell[0] + '.' + otch_cell[0], alfa_dic)

            # Должность
            elif xl_col == 3:
                dolz_cell = cell_value

            # Подразделение
            elif xl_col == 4:
                podr_cell = cell_value

            # Роль
            elif xl_col == 6:
                role_cell_list = []
                if 'руководител' in cell_value.lower():
                    role_cell_list.append('roles.gasps.79997554')  # Рук со/од
                if 'начальник' in cell_value.lower():
                    role_cell_list.append('roles.gasps.79997554')  # Рук со/од
                if ('дознаватель' in cell_value.lower()) or ('дознователь' in cell_value.lower()):
                    role_cell_list.append('roles.gasps.79997634')  # Сл/До
                if 'следовател' in cell_value.lower():
                    role_cell_list.append('roles.gasps.79997634')  # Сл/До
                if 'оператор' in cell_value.lower():
                    role_cell_list.append('roles.gasps.80076916')  # Оператор
                if 'аналитик' in cell_value.lower():
                    role_cell_list.append('roles.gasps.80076855')  # Аналитик
                if not role_cell_list:
                    role_cell_list.append('roles.gasps.80076855')  # Аналитик
                role_cell = str(role_cell_list).replace("'", '')

            # Отпечаток сертификата
            elif xl_col == 8:
                cell_value = stripper(cell_value).upper()
                if cell_value not in otp_set_exists:
                    otp_list.append(cell_value)
                    if otp_list.count(cell_value) == 1 and len(cell_value) == otp_len:
                        otp_cell = cell_value
                    else:
                        otp_cell = None
                else:
                    otp_cell = None

        # Пароль
        pass_cell = 'Qwerty1'

        if flag_add_in_new_file:
            # собираю строку по правилу выгрузки и добавляю её в файл
            wb_new_s.append([user_name, fam_cell, imya_cell, otch_cell, dolz_cell,
                             pass_cell, otp_cell, podr_cell, role_cell])

            print(user_name, fam_cell, imya_cell, otch_cell, dolz_cell,
                  pass_cell, otp_cell, podr_cell, role_cell, sep=', ')
        else:
            print(f' ... !!! пользователь существует - {fam_cell} {imya_cell} {otch_cell}, и не будет добавлен')

# сохраняю итоговый файл в эксель
wb_new.save(xl_f_to)
wb_from.close()
wb_new.close()

# подсчёт времени
time_finish = time.monotonic()
print('................закончено за', round(time_finish - time_start, 3), 'секунд')

# закрываю программу
input('Нажмите ENTER')
